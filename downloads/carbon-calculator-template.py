#!/usr/bin/env python3
"""
Event Carbon Footprint Calculator - Excel Template Generator
Generates an Excel spreadsheet with embedded formulas and emissions factors.
Run: python3 carbon-calculator-template.py
Output: carbon-event-template.xlsx
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Install with: pip install openpyxl")
    exit(1)

# Emissions Factors (must match the web calculator)
MEAL_FACTORS = {
    'Vegan': 0.39,
    'Vegetarian': 0.51,
    'White Meat': 1.58,
    'Red Meat': 7.26,
    'Pescetarian': 1.55,
}

COFFEE_FACTOR = 0.25

TRANSPORT_FACTORS = {
    'Walking': 0,
    'Bicycle': 5,
    'Car (solo)': 215.6,
    'Carpool': 107.8,
    'Ferry': 138,
    'Train (TGV)': 33.3,
    'Train (Western Europe)': 16,
    'Flight (short <1000km)': 258.6,
    'Flight (medium 1000-3000km)': 107.5,
    'Flight (long >3000km)': 152,
}

def create_excel_template():
    """Create Excel workbook with carbon calculator template."""
    wb = Workbook()
    
    # Remove default sheet and create new ones
    wb.remove(wb.active)
    
    # ====================================================================
    # SHEET 1: Calculator (Main)
    # ====================================================================
    ws_calc = wb.create_sheet("Calculator", 0)
    
    # Styles
    header_fill = PatternFill(start_color="2D5B3F", end_color="2D5B3F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    section_fill = PatternFill(start_color="E8B859", end_color="E8B859", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF", size=11)
    
    input_fill = PatternFill(start_color="FFFBF5", end_color="FFFBF5", fill_type="solid")
    result_fill = PatternFill(start_color="F0F4F1", end_color="F0F4F1", fill_type="solid")
    
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Title
    ws_calc['A1'] = "🌍 Event Carbon Footprint Calculator"
    ws_calc['A1'].font = Font(bold=True, size=16, color="2D5B3F")
    ws_calc.merge_cells('A1:D1')
    
    ws_calc['A2'] = "Enter your event details below. All calculations happen automatically."
    ws_calc.merge_cells('A2:D2')
    ws_calc['A2'].font = Font(italic=True, size=10, color="888888")
    
    row = 4
    
    # EVENT OVERVIEW
    ws_calc[f'A{row}'] = "EVENT OVERVIEW"
    ws_calc[f'A{row}'].fill = section_fill
    ws_calc[f'A{row}'].font = section_font
    ws_calc.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws_calc[f'A{row}'] = "Number of Attendees"
    ws_calc[f'B{row}'] = 100  # Default value
    ws_calc[f'B{row}'].fill = input_fill
    ws_calc[f'C{row}'] = "(edit this)"
    ws_calc[f'C{row}'].font = Font(italic=True, size=9, color="888888")
    row += 2
    
    # MEALS SECTION
    ws_calc[f'A{row}'] = "MEALS & REFRESHMENTS"
    ws_calc[f'A{row}'].fill = section_fill
    ws_calc[f'A{row}'].font = section_font
    ws_calc.merge_cells(f'A{row}:D{row}')
    row += 1
    
    # Meal headers
    ws_calc[f'A{row}'] = "Category"
    ws_calc[f'B{row}'] = "Count"
    ws_calc[f'C{row}'] = "Factor (kg CO2e)"
    ws_calc[f'D{row}'] = "Total (kg CO2e)"
    for col in ['A', 'B', 'C', 'D']:
        ws_calc[f'{col}{row}'].fill = result_fill
        ws_calc[f'{col}{row}'].font = Font(bold=True, size=10)
        ws_calc[f'{col}{row}'].border = border
    row += 1
    
    meal_row_start = row
    for meal_type, factor in MEAL_FACTORS.items():
        ws_calc[f'A{row}'] = meal_type
        ws_calc[f'B{row}'] = 0  # Input: user fills this
        ws_calc[f'B{row}'].fill = input_fill
        ws_calc[f'C{row}'] = factor
        ws_calc[f'D{row}'] = f"=B{row}*C{row}"  # Formula
        for col in ['A', 'B', 'C', 'D']:
            ws_calc[f'{col}{row}'].border = border
        row += 1
    
    meal_row_end = row - 1
    
    # Coffee
    ws_calc[f'A{row}'] = "Coffee/Tea Breaks"
    ws_calc[f'B{row}'] = 0
    ws_calc[f'B{row}'].fill = input_fill
    ws_calc[f'C{row}'] = COFFEE_FACTOR
    ws_calc[f'D{row}'] = f"=B{row}*C{row}"
    for col in ['A', 'B', 'C', 'D']:
        ws_calc[f'{col}{row}'].border = border
    coffee_row = row
    row += 2
    
    # Meals Total
    ws_calc[f'A{row}'] = "TOTAL MEALS"
    ws_calc[f'A{row}'].font = Font(bold=True)
    ws_calc[f'D{row}'] = f"=SUM(D{meal_row_start}:D{meal_row_end},D{coffee_row})"
    ws_calc[f'D{row}'].fill = PatternFill(start_color="E8ECEA", end_color="E8ECEA", fill_type="solid")
    ws_calc[f'D{row}'].font = Font(bold=True, size=11)
    meals_total_row = row
    row += 2
    
    # TRANSPORT SECTION
    ws_calc[f'A{row}'] = "TRANSPORTATION"
    ws_calc[f'A{row}'].fill = section_fill
    ws_calc[f'A{row}'].font = section_font
    ws_calc.merge_cells(f'A{row}:E{row}')
    row += 1
    
    ws_calc[f'A{row}'] = "People"
    ws_calc[f'B{row}'] = "Distance (km)"
    ws_calc[f'C{row}'] = "Mode"
    ws_calc[f'D{row}'] = "Factor (g CO2e/km)"
    ws_calc[f'E{row}'] = "Total (kg CO2e)"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_calc[f'{col}{row}'].fill = result_fill
        ws_calc[f'{col}{row}'].font = Font(bold=True, size=10)
        ws_calc[f'{col}{row}'].border = border
    row += 1
    
    transport_row_start = row
    # Add 5 empty journey rows
    for i in range(5):
        ws_calc[f'A{row}'] = 1
        ws_calc[f'A{row}'].fill = input_fill
        ws_calc[f'B{row}'] = 0
        ws_calc[f'B{row}'].fill = input_fill
        ws_calc[f'C{row}'] = ""
        ws_calc[f'C{row}'].fill = input_fill
        ws_calc[f'D{row}'] = 0
        ws_calc[f'D{row}'].fill = input_fill
        ws_calc[f'E{row}'] = f"=IF(B{row}=0,0,A{row}*B{row}*D{row}/1000)"
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_calc[f'{col}{row}'].border = border
        row += 1
    
    transport_row_end = row - 1
    
    row += 1
    ws_calc[f'A{row}'] = "TOTAL TRANSPORT"
    ws_calc[f'A{row}'].font = Font(bold=True)
    ws_calc[f'E{row}'] = f"=SUM(E{transport_row_start}:E{transport_row_end})"
    ws_calc[f'E{row}'].fill = PatternFill(start_color="E8ECEA", end_color="E8ECEA", fill_type="solid")
    ws_calc[f'E{row}'].font = Font(bold=True, size=11)
    transport_total_row = row
    row += 3
    
    # SUMMARY
    ws_calc[f'A{row}'] = "SUMMARY"
    ws_calc[f'A{row}'].fill = header_fill
    ws_calc[f'A{row}'].font = header_font
    ws_calc.merge_cells(f'A{row}:D{row}')
    row += 1
    
    ws_calc[f'A{row}'] = "Total Emissions (kg CO2e)"
    ws_calc[f'B{row}'] = f"=D{meals_total_row}+E{transport_total_row}"
    ws_calc[f'B{row}'].fill = result_fill
    ws_calc[f'B{row}'].font = Font(bold=True, size=11)
    total_emissions_row = row
    row += 1
    
    ws_calc[f'A{row}'] = "Per Attendee (kg CO2e)"
    ws_calc[f'B{row}'] = f"=IF(B4=0,0,B{total_emissions_row}/B4)"
    ws_calc[f'B{row}'].fill = result_fill
    ws_calc[f'B{row}'].font = Font(bold=True, size=11)
    row += 1
    
    ws_calc[f'A{row}'] = "Meals Impact (kg CO2e)"
    ws_calc[f'B{row}'] = f"=D{meals_total_row}"
    ws_calc[f'B{row}'].fill = result_fill
    row += 1
    
    ws_calc[f'A{row}'] = "Transport Impact (kg CO2e)"
    ws_calc[f'B{row}'] = f"=E{transport_total_row}"
    ws_calc[f'B{row}'].fill = result_fill
    row += 2
    
    # Column widths
    ws_calc.column_dimensions['A'].width = 25
    ws_calc.column_dimensions['B'].width = 18
    ws_calc.column_dimensions['C'].width = 20
    ws_calc.column_dimensions['D'].width = 20
    ws_calc.column_dimensions['E'].width = 20
    
    # ====================================================================
    # SHEET 2: Factors (Reference)
    # ====================================================================
    ws_factors = wb.create_sheet("Factors", 1)
    
    row = 1
    ws_factors[f'A{row}'] = "EMISSIONS FACTORS"
    ws_factors[f'A{row}'].font = Font(bold=True, size=14, color="2D5B3F")
    row += 2
    
    # Meals factors
    ws_factors[f'A{row}'] = "MEALS (kg CO2e per meal)"
    ws_factors[f'A{row}'].fill = section_fill
    ws_factors[f'A{row}'].font = section_font
    ws_factors.merge_cells(f'A{row}:B{row}')
    row += 1
    
    for meal_type, factor in MEAL_FACTORS.items():
        ws_factors[f'A{row}'] = meal_type
        ws_factors[f'B{row}'] = factor
        row += 1
    
    ws_factors[f'A{row}'] = "Coffee/Tea"
    ws_factors[f'B{row}'] = COFFEE_FACTOR
    row += 2
    
    # Transport factors
    ws_factors[f'A{row}'] = "TRANSPORT (g CO2e per km)"
    ws_factors[f'A{row}'].fill = section_fill
    ws_factors[f'A{row}'].font = section_font
    ws_factors.merge_cells(f'A{row}:B{row}')
    row += 1
    
    for mode, factor in TRANSPORT_FACTORS.items():
        ws_factors[f'A{row}'] = mode
        ws_factors[f'B{row}'] = factor
        row += 1
    
    row += 2
    ws_factors[f'A{row}'] = "SOURCES"
    ws_factors[f'A{row}'].font = Font(bold=True, size=11)
    row += 1
    
    sources = [
        "Labos1.5",
    ]
    
    for source in sources:
        ws_factors[f'A{row}'] = source
        ws_factors[f'A{row}'].font = Font(italic=True, size=9, color="555555")
        ws_factors.merge_cells(f'A{row}:B{row}')
        row += 1
    
    ws_factors.column_dimensions['A'].width = 50
    ws_factors.column_dimensions['B'].width = 20
    
    # Save
    filename = 'carbon-event-template.xlsx'
    wb.save(filename)
    print(f"✓ Excel template created: {filename}")
    print(f"  - Sheet 1: Calculator (enter your data)")
    print(f"  - Sheet 2: Factors (reference & sources)")
    print()
    print("How to use:")
    print("  1. Open carbon-event-template.xlsx in Excel, Calc, or Google Sheets")
    print("  2. Edit the 'Calculator' sheet:")
    print("     - B4: Number of attendees")
    print("     - Column B (Meals): Count for each meal type")
    print("     - Rows 16-20 (Journeys): Fill people, distance, and mode")
    print("  3. All totals calculate automatically")
    print()
    print("Note: This template matches the web calculator's factors exactly.")
    print("      For the most up-to-date calculator, use: carbon-calculator.html")

if __name__ == '__main__':
    create_excel_template()
